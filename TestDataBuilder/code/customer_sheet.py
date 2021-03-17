#!/usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import datetime
import common
import china_regions_dict as CRD


headlist = ['客户名称', '手机号', '归属用户', '归属渠道', '客户证件类型', '客户证件号码',
            '车牌号', '车辆识别代号', '车型名称', '行驶城市', '车辆注册日期', '保险公司',
            '商业险到期日', '交强险到期日', '商业险保单号', '交强险保单号', '客户标签', '备注'
            ]

# inwb = load_workbook('客户test1.xlsx')
wb = Workbook()
# 激活 worksheet
ws = wb.active
# headlist分配到第一行中
ws.append(headlist)
# 商业险到期时间=交强险到期时间
instime = mustime = common.endTime()
province_city_district = CRD.regions()
# 表格内容infolist
infolist = [common.customName(), common.customPhone(), '14100000003',
            '飞侠太保杭州', '身份证', common.customID(), common.carPlate(),
            common.vinNum(), '东风标致DC7164DB轿车', province_city_district[1],
            common.regTime(), common.companyName(),
            instime, mustime, common.policyNum(), common.policyNum(), common.customTag(), common.remarks(10)]

# 附加infolist，从第一行开始附加x条数据
if __name__ == '__main__':
    x = 10
    # NUM = input('需数据n条：')
    # print('打印%s条数据...' % NUM)
    # x = int(NUM)
    for num in range(x):
        ws.append(infolist)
        print(infolist)
    # input('Press Enter to exit...')

# 保存文件
CUS_FILE = '客户test' + datetime.datetime.now().strftime('%Y%m%d%H%M%S') + '.xlsx'
wb.save(CUS_FILE)

# BASE_DIR = os.path.dirname(__file__)
# DC_PATH = BASE_DIR + '\\' + datetime.datetime.now().strftime('%Y%m%d') + '.xlsx'
# 获取到当前文件的目录，并检查是否有该文件，如果不存在则自动新建文件
# if not os.path.exists(DC_PATH):
#     open(DC_PATH, 'w')
